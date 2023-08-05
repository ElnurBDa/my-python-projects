from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl import worksheet

wb = load_workbook('file_X.xlsx')
ws = wb['Main']

cols=[]
for x in range(1,5):
    rows=[]
    for row in ws.iter_rows(min_row=1, max_col=x+1, max_row=41, values_only=True):
        rows.append((row[0],row[x]))
    cols.append(rows)
for x in cols:
    print(x)
    ws=wb.create_sheet(x[0][1])
    ws.append(x[0])
    for y in x[1::]:
        if y[1]>=5:
            ws.append(y)










wb.save('file_X.xlsx')



